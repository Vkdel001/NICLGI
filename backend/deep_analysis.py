import pandas as pd
import os
from openpyxl import load_workbook

backend_file = r'C:\Users\Ryan EZ\OneDrive - EZ DASH LTD\Documents\letters_Motors_health\backend\RENEWAL_LISTING.xlsx'

print("=== DEEP ANALYSIS OF RENEWAL_LISTING.xlsx ===")
print(f"File: {backend_file}")
print(f"File exists: {os.path.exists(backend_file)}")

if os.path.exists(backend_file):
    print(f"File size: {os.path.getsize(backend_file)} bytes")
    
    # Check with openpyxl first
    print("\n--- OPENPYXL ANALYSIS ---")
    try:
        wb = load_workbook(backend_file)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Sheet '{sheet_name}': {ws.max_row} rows, {ws.max_column} columns")
            
            # Check if there's data in the sheet
            data_rows = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data_rows += 1
            print(f"Sheet '{sheet_name}': {data_rows} rows with data")
            
    except Exception as e:
        print(f"Openpyxl error: {e}")
    
    # Check with pandas - default behavior
    print("\n--- PANDAS DEFAULT READ ---")
    try:
        df_default = pd.read_excel(backend_file)
        print(f"Pandas default: {len(df_default)} rows")
        if 'POL_NO' in df_default.columns:
            valid_policies = df_default['POL_NO'].notna().sum()
            print(f"Valid policies: {valid_policies}")
        print(f"Columns: {list(df_default.columns)}")
    except Exception as e:
        print(f"Pandas default error: {e}")
    
    # Check each sheet individually with pandas
    print("\n--- PANDAS SHEET-BY-SHEET ---")
    try:
        excel_file = pd.ExcelFile(backend_file)
        print(f"Sheet names from pandas: {excel_file.sheet_names}")
        
        for sheet_name in excel_file.sheet_names:
            df_sheet = pd.read_excel(backend_file, sheet_name=sheet_name)
            print(f"Sheet '{sheet_name}': {len(df_sheet)} rows")
            if 'POL_NO' in df_sheet.columns:
                valid_policies = df_sheet['POL_NO'].notna().sum()
                print(f"  Valid policies in '{sheet_name}': {valid_policies}")
                
                # Show first few policy numbers
                valid_df = df_sheet[df_sheet['POL_NO'].notna()]
                if len(valid_df) > 0:
                    print(f"  First 3 policies: {valid_df['POL_NO'].head(3).tolist()}")
                    
    except Exception as e:
        print(f"Pandas sheet analysis error: {e}")
    
    # Check with different engines
    print("\n--- DIFFERENT ENGINES ---")
    engines = ['openpyxl', 'xlrd']
    for engine in engines:
        try:
            df_engine = pd.read_excel(backend_file, engine=engine)
            print(f"Engine '{engine}': {len(df_engine)} rows")
            if 'POL_NO' in df_engine.columns:
                valid_policies = df_engine['POL_NO'].notna().sum()
                print(f"  Valid policies with '{engine}': {valid_policies}")
        except Exception as e:
            print(f"Engine '{engine}' error: {e}")

else:
    print("File not found!")

# Also check what the healthcare script is actually reading
print("\n=== HEALTHCARE SCRIPT SIMULATION ===")
print("Simulating what healthcare_renewal_final.py reads...")

# Change to backend directory (like the script does)
original_dir = os.getcwd()
try:
    backend_dir = r'C:\Users\Ryan EZ\OneDrive - EZ DASH LTD\Documents\letters_Motors_health\backend'
    os.chdir(backend_dir)
    print(f"Changed to directory: {os.getcwd()}")
    
    # Try to read RENEWAL_LISTING.xlsx from current directory
    if os.path.exists("RENEWAL_LISTING.xlsx"):
        print("Found RENEWAL_LISTING.xlsx in backend directory")
        df_script = pd.read_excel("RENEWAL_LISTING.xlsx", engine='openpyxl')
        print(f"Script would read: {len(df_script)} rows")
        if 'POL_NO' in df_script.columns:
            valid_policies = df_script['POL_NO'].notna().sum()
            print(f"Script would process: {valid_policies} valid policies")
    else:
        print("RENEWAL_LISTING.xlsx not found in backend directory")
        
finally:
    os.chdir(original_dir)